from __future__ import annotations

# ========== FIX IMPORTS =============

import pandas as pd
from openpyxl import load_workbook
from .ADLSClient import ADLSClient
from utils.constants import *
from st_web.utils.helpers import get_asset_file_path
from __future__ import annotations

import os
import pyspark.pandas as ps
from io import BytesIO
from databricks.sdk.runtime import dbutils
from utils.constants import *
from dbc.utils.helpers import get_filename_from_folder

# Create ADLS connection object
adls_client = ADLSClient()

# Get env variables
ADLS_CONTAINER = os.getenv('ADLS_CONTAINER')
DBFS_TMP_PATH = os.getenv('DBFS_TMP_PATH')

# ===== FROM ST_WEB =====
# Read text file from assets folder
def get_asset_text(text_file_name: str) -> str:
    
    # Find text path
    text_path = get_asset_file_path(ASSET_TYPE_TEXTS, f'{text_file_name}{TEXT_FILE_EXTENSION}')

    # Read sales uploader text
    with open(text_path, 'r', encoding='utf-8') as f:
        # Read file content into text variable
        text = f.read()

    # Return the text content
    return text

# Get the name of a folder's only file
def get_filename_from_folder(path:str) -> str:

    # Return the first file name in the folder
    for file_info in dbutils.fs.ls(path):
        return file_info.name

# ===== FROM DBC =====
# # Read Excel sheet from ADLS layer and return as openpyxl workbook
# def read_sheet(layer: str, year: str, month: str):
#     from openpyxl import load_workbook
    
#     # Construct the file path using layer, year, and month
#     path = f'{ADLS_CONTAINER}{layer}/{ADLS_CATEGORY_SALES}/{year}/{month}'

#     # Get desired sheet's name
#     file_name = get_filename_from_folder(path)

#     # Create tmp folder
#     dbutils.fs.mkdirs(DBFS_TMP_PATH)

#     # Copy sheet to tmp folder
#     dbutils.fs.cp(
#         f'{path}/{file_name}',
#         DBFS_TMP_PATH
#     )

#     tmp_sheet_path = f'../tmp/{file_name}'

#     # Return the openpyxl workbook
#     return load_workbook(tmp_sheet_path)

# ===== FROM ST_WEB =====
# Read Excel sheet from ADLS layer and return as openpyxl workbook
def read_sheet(layer: str, year: str, month: str):

    # Construct the file path using layer, year, and month
    path = f'{layer}/{ADLS_CATEGORY_SALES}/{year}/{month}'

    # Read the Excel file from ADLS as bytes
    sheet_bytes = adls_client.read_folder(path)

    # Return the openpyxl workbook
    return load_workbook(sheet_bytes)

# ===== FROM ST_WEB =====
# Read tabular file from ADLS layer and return as DataFrame
def read_tabular(layer: str, category: str, year: str, month: str):

    # Construct the file path using layer, category, year, and month
    path = f'{layer}/{category}/{year}/{month}'

    # Read the parquet file from ADLS as bytes
    parquet_bytes = adls_client.read_folder(path)

    # Return the parquet DataFrame
    return pd.read_parquet(parquet_bytes)